import tkinter as tk
from tkinter import ttk, messagebox, filedialog, Scrollbar
import pandas as pd
from faker import Faker
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import os
import subprocess

fake = Faker('zh_CN')


class DataGeneratorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("数据随机生成器")
        self.root.geometry("440x620")

        # 可选字段
        self.available_fields = ["姓名", "性别", "身份证号", "住址", "联系方式"]
        self.selected_fields = []

        # 数量设置
        self.num_entries_var = tk.IntVar()
        self.num_entries_var.set(1)

        # 固定下载格式为 Excel
        self.download_format = "Excel"

        # 新增姓氏输入框
        self.surname_var = tk.StringVar()
        self.village_name_var = tk.StringVar()
        self.region_code_var = tk.StringVar()

        # 创建界面组件
        self.create_widgets()

    def create_widgets(self):
        # 姓氏输入框
        self.surname_label = tk.Label(self.root, text="请输入姓氏（多个用英文逗号分隔）:")
        self.surname_entry = tk.Entry(self.root, textvariable=self.surname_var)
        self.surname_label.grid(row=4, column=0, columnspan=3, padx=10, pady=5, sticky=tk.W)
        self.surname_entry.grid(row=5, column=0, columnspan=3, padx=10, pady=5, sticky=tk.W + tk.E)
        self.surname_label.grid_remove()
        self.surname_entry.grid_remove()

        # 地区编码输入框
        self.region_code_label = tk.Label(self.root, text="请输入地区编码（身份证前6位）:")
        self.region_code_entry = tk.Entry(self.root, textvariable=self.region_code_var)
        self.region_code_label.grid(row=8, column=0, columnspan=3, padx=10, pady=5, sticky=tk.W)
        self.region_code_entry.grid(row=9, column=0, columnspan=3, padx=10, pady=5, sticky=tk.W + tk.E)
        self.region_code_label.grid_remove()
        self.region_code_entry.grid_remove()

        # 村庄名称输入框
        self.village_name_label = tk.Label(self.root, text="请输入村庄名称:")
        self.village_name_entry = tk.Entry(self.root, textvariable=self.village_name_var)
        self.village_name_label.grid(row=6, column=0, columnspan=3, padx=10, pady=5, sticky=tk.W)
        self.village_name_entry.grid(row=7, column=0, columnspan=3, padx=10, pady=5, sticky=tk.W + tk.E)
        self.village_name_label.grid_remove()
        self.village_name_entry.grid_remove()

        # 字段选择下拉框
        self.field_combobox = ttk.Combobox(self.root, values=self.available_fields)
        self.field_combobox.grid(row=0, column=0, padx=10, pady=10)

        # 添加和删除字段按钮
        add_button = tk.Button(self.root, text="添加字段", command=self.add_field)
        add_button.grid(row=0, column=1, padx=1, pady=10)

        delete_button = tk.Button(self.root, text="删除字段", command=self.delete_field)
        delete_button.grid(row=0, column=2, padx=1, pady=10)

        # 显示已选字段
        self.selected_fields_listbox = tk.Listbox(self.root, width=60)
        self.selected_fields_listbox.grid(row=1, column=0, columnspan=3, pady=8)

        # 数量设置
        num_entries_label = tk.Label(self.root, text="生成数据条数:")
        num_entries_label.grid(row=2, column=0, padx=10, pady=5)
        num_entries_entry = tk.Entry(self.root, textvariable=self.num_entries_var)
        num_entries_entry.grid(row=2, column=1, padx=10, pady=5)

        # 生成数据按钮
        generate_button = tk.Button(self.root, text="生成数据", command=self.generate_data)
        generate_button.grid(row=3, column=0, columnspan=3, pady=20)

    def add_field(self):
        field = self.field_combobox.get()
        if field and field not in self.selected_fields:
            self.selected_fields.append(field)
            self.selected_fields_listbox.insert(tk.END, field)

            if field == "姓名":
                self.surname_label.grid()
                self.surname_entry.grid()
            elif field == "住址":
                self.village_name_label.grid()
                self.village_name_entry.grid()
            elif field == "身份证号":
                self.region_code_label.grid()
                self.region_code_entry.grid()

    def delete_field(self):
        selected_index = self.selected_fields_listbox.curselection()
        if selected_index:
            index = selected_index[0]
            field = self.selected_fields_listbox.get(index)
            self.selected_fields.remove(field)
            self.selected_fields_listbox.delete(index)

            if field == "姓名":
                self.surname_label.grid_remove()
                self.surname_entry.grid_remove()
            elif field == "住址":
                self.village_name_label.grid_remove()
                self.village_name_entry.grid_remove()
            elif field == "身份证号":
                self.region_code_label.grid_remove()
                self.region_code_entry.grid_remove()

    def generate_data(self):
        if not self.selected_fields:
            messagebox.showwarning("警告", "请至少选择一个字段。")
            return

        if "姓名" in self.selected_fields and not self.surname_var.get():
            messagebox.showwarning("警告", "请输入姓氏。")
            return
        if "身份证号" in self.selected_fields and not self.region_code_var.get():
            messagebox.showwarning("警告", "请输入地区编码。")
            return
        if "住址" in self.selected_fields and not self.village_name_var.get():
            messagebox.showwarning("警告", "请输入村庄名称。")
            return

        surnames = [s.strip() for s in self.surname_var.get().split(',')]
        num_entries = self.num_entries_var.get()
        data = []
        used_names = set()  # 用于存储已生成的姓名
        for _ in range(num_entries):
            entry = {}
            for field in self.selected_fields:
                if field == "姓名":
                    name = None
                    while True:
                        chosen_surname = fake.random_element(surnames)
                        # 80% 生成男性名字
                        if fake.random.random() < 0.8:
                            given_name = fake.first_name_male()
                        else:
                            given_name = fake.first_name_female()
                        # 90% 生成三字名字
                        if fake.random.random() < 0.9:
                            second_given_name = fake.first_name()
                            given_name = given_name + second_given_name
                        name = chosen_surname + given_name
                        # 确保姓名不超过 3 个字
                        if len(name) <= 3:
                            if name not in used_names:
                                used_names.add(name)
                                break
                    entry[field] = name
                elif field == "性别":
                    entry[field] = fake.random_element(elements=("男", "女"))
                elif field == "身份证号":
                    birth_date = fake.date_of_birth(minimum_age=18, maximum_age=80)
                    birth_date_str = birth_date.strftime("%Y%m%d")
                    sequence_num = fake.random_int(min=100, max=999)
                    id_number = f"{self.region_code_var.get()}{birth_date_str}{sequence_num}"
                    # 简单的校验码计算（实际更复杂）
                    factors = [7, 9, 10, 5, 8, 4, 2, 1, 6, 3, 7, 9, 10, 5, 8, 4, 2]
                    check_code_list = ['1', '0', 'X', '9', '8', '7', '6', '5', '4', '3', '2']
                    total = 0
                    for i in range(17):
                        total += int(id_number[i]) * factors[i]
                    check_code = check_code_list[total % 11]
                    id_number += check_code
                    entry[field] = id_number
                elif field == "住址":
                    # 简化地址生成
                    village = self.village_name_var.get()
                    street_num = fake.random_int(min=1, max=100)
                    entry[field] = f"{village} {street_num}号"
                elif field == "联系方式":
                    entry[field] = fake.phone_number()
            data.append(entry)

        df = pd.DataFrame(data)
        # 按照 selected_fields 列表的顺序排列 DataFrame 的列
        df = df[self.selected_fields]

        if df.empty:
            messagebox.showwarning("警告", "生成的数据为空，请检查设置。")
            return

        # 预览功能
        preview_window = tk.Toplevel(self.root)
        preview_window.title("数据预览")
        preview_window.configure(bg="#f0f0f0")

        # 下载按钮
        download_button = tk.Button(preview_window, text="下载数据", command=lambda: self.download_data(df), font=("Arial", 12), bg="#4CAF50", fg="white")
        download_button.grid(row=0, column=0, padx=10, pady=10, sticky=tk.NW)

        # 创建一个画布用于放置表格和滚动条
        canvas = tk.Canvas(preview_window, bg="#f0f0f0")
        canvas.grid(row=1, column=0, columnspan=2, sticky=tk.NSEW)
        preview_window.grid_rowconfigure(1, weight=1)
        preview_window.grid_columnconfigure(0, weight=1)

        # 创建垂直滚动条
        v_scrollbar = Scrollbar(preview_window, command=canvas.yview)
        v_scrollbar.grid(row=1, column=1, sticky=tk.NS)
        canvas.configure(yscrollcommand=v_scrollbar.set)

        # 创建一个框架用于放置表格
        table_frame = tk.Frame(canvas, bg="white", bd=1, relief=tk.SOLID)
        canvas.create_window((0, 0), window=table_frame, anchor=tk.NW)

        # 计算每列的最大宽度
        col_widths = [0] * len(df.columns)
        for i, col in enumerate(df.columns):
            col_widths[i] = max(col_widths[i], len(str(col)))
        for row_index, row in df.iterrows():
            for col_index, value in enumerate(row):
                col_widths[col_index] = max(col_widths[col_index], len(str(value)))

        # 显示表头
        for i, col in enumerate(df.columns):
            label = tk.Label(table_frame, text=col, font=("Arial", 12, "bold"), bg="#e0e0e0", fg="black", bd=1, relief=tk.SOLID, padx=10, pady=5)
            label.grid(row=0, column=i, sticky="nsew")
            table_frame.columnconfigure(i, minsize=col_widths[i] * 8)

        # 显示数据
        for row_index, row in df.iterrows():
            for col_index, value in enumerate(row):
                entry = tk.Label(table_frame, text=value, font=("Arial", 10), bg="white", bd=1, relief=tk.SOLID, justify=tk.LEFT, anchor="w")
                entry.grid(row=row_index + 1, column=col_index, sticky="nsew")

        # 设置行的权重，使行能自适应高度
        for i in range(len(df) + 1):
            table_frame.rowconfigure(i, weight=1)

        # 更新画布的滚动区域
        table_frame.update_idletasks()
        canvas.config(scrollregion=canvas.bbox("all"))

        # 计算表格的宽度和高度
        table_width = table_frame.winfo_reqwidth()
        table_height = table_frame.winfo_reqheight()

        # 考虑滚动条和按钮的空间
        total_width = table_width + v_scrollbar.winfo_reqwidth()
        total_height = table_height + download_button.winfo_reqheight() + 20

        # 设置窗口大小
        preview_window.geometry(f"{total_width}x{total_height}")

    def download_data(self, df):
        file_extension = ".xlsx"

        file_path = filedialog.asksaveasfilename(defaultextension=file_extension,
                                                 filetypes=[("Excel Files", "*.xlsx")])
        if file_path:
            try:
                wb = Workbook()
                ws = wb.active

                # 写入表头
                header = df.columns.tolist()
                ws.append(header)
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # 写入数据
                for row in df.values.tolist():
                    ws.append(row)

                # 自动调整列宽
                for col_num, column in enumerate(ws.columns, 1):
                    max_length = 0
                    column_letter = get_column_letter(col_num)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column_letter].width = adjusted_width

                wb.save(file_path)

                messagebox.showinfo("提示", f"数据已成功下载到 {file_path}。")

                # 根据操作系统打开文件
                if os.name == "nt":
                    os.startfile(file_path)
                elif os.name == "posix":
                    if os.system("which open") == 0:
                        subprocess.call(["open", file_path])
                    elif os.system("which xdg-open") == 0:
                        subprocess.call(["xdg-open", file_path])
                    else:
                        messagebox.showwarning("提示", "无法找到合适的程序打开文件。")

            except Exception as e:
                messagebox.showerror("错误", f"保存文件时出现错误: {str(e)}")


if __name__ == "__main__":
    root = tk.Tk()
    app = DataGeneratorApp(root)
    root.mainloop()